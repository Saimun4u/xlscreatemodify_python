import openpyxl
import os

wb = openpyxl.Workbook()

sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'] = 'Name'

sheet['A2'] = 'Saimun'

sheet['B1'] = 'Passion'

sheet['B2'] = 'Coding'

sheet2 = wb.create_sheet()

travel = wb.create_sheet('travel')

wb.create_sheet(index=0, title='savings')

print(wb.get_sheet_names())

os.chdir('C:\\Users\Saimun\Python ardit\Editing excel spreadsheets')

wb.save('example.xlsx')

